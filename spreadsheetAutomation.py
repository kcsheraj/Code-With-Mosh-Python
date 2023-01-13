import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#method takes file and caluclates 90 percent of the thrid col and put it in next col
#and also creates bar chart
#use 'transactions.xlsx'
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2,sheet.max_row+1):#get each cell in col 3
        cell = sheet.cell(row,3) #does same as sheet['A1]
        corrected_price = cell.value * 0.10
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    #slect values in 4th col
    values = Reference(sheet,
            min_row=2,#select cells from row 2 to 4
            max_row=4,
            min_col=4,#slect cells from col 4 to 4
            max_col=4,
            )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')#add chart at e2

    wb.save('transactions2.xlsx')#save changes in new file

process_workbook('transactions.xlsx')